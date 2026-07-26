from __future__ import annotations

import csv
import hashlib
import re
import sqlite3
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Callable, Iterable

import openpyxl

from .exercises import get_exercise_offset_days, scheduled_date_from

WEEK_HEADER_RE = re.compile(r"^SEMANA\s*(\d{1,2})$", re.IGNORECASE)
LESSON_RE = re.compile(r"^(?P<subject_module>.+?)\s*-\s*Aula\s*(?P<lesson_number>\d+)\s*$", re.IGNORECASE)
PENDING_RE = re.compile(r"^(?P<subject>.+?)\s+UFPR\s*\(aguardando edital\)\s*$", re.IGNORECASE)
ROMAN_RE = re.compile(r"^(?P<subject>.+?)\s+(?P<module>[IVXLCDM]+)$", re.IGNORECASE)

DAY_COLUMNS = [
    (2, 1, "Segunda"),
    (4, 2, "Terça"),
    (6, 3, "Quarta"),
    (8, 4, "Quinta"),
    (10, 5, "Sexta"),
    (12, 6, "Sábado"),
    (14, 7, "Domingo"),
]

SUBJECT_PREFIXES = {
    "biologia": "BIO",
    "fisica": "FIS",
    "filosofia": "FIL",
    "geografia": "GEO",
    "historia": "HIS",
    "lingua estrangeira": "LIN",
    "literatura": "LIT",
    "matematica": "MAT",
    "portugues": "POR",
    "quimica": "QUI",
    "redacao": "RED",
    "sociologia": "SOC",
}

PORTUGUESE_DAY_NAMES = {
    1: "Segunda",
    2: "Terça",
    3: "Quarta",
    4: "Quinta",
    5: "Sexta",
    6: "Sábado",
    7: "Domingo",
}

DEFAULT_UN_START_DATE = date(2026, 3, 13)
FO_METADATA_CSV_PATH = (
    Path(__file__).resolve().parent.parent / "work" / "fo_bridge" / "output" / "fo_all_lessons_metadata.csv"
)
UN_DURATION_CSV_PATH = (
    Path(__file__).resolve().parent.parent / "work" / "gpe_bridge" / "output" / "un_video_durations.csv"
)


@dataclass
class ParsedLesson:
    slot_key: str
    lesson_code: str
    track_code: str
    lesson_type: str
    title_raw: str
    portal_title: str | None
    relative_path: str | None
    external_url: str | None
    duration_seconds: int | None
    subject_name: str | None
    subject_prefix: str | None
    module_label: str | None
    module_number: int | None
    lesson_number: int | None
    week_number: int
    day_index: int
    day_name: str
    slot_index: int
    recommended_date: str
    source_sheet: str


@dataclass(frozen=True)
class FoImportPreflight:
    incoming_count: int
    existing_fo_count: int
    matching_lesson_codes: tuple[str, ...]
    new_lesson_codes: tuple[str, ...]
    missing_lesson_codes: tuple[str, ...]
    errors: tuple[str, ...]

    @property
    def is_safe(self) -> bool:
        return not self.errors

    def require_safe(self) -> None:
        if self.is_safe:
            return
        details = "\n".join(f"- {message}" for message in self.errors)
        raise FoImportPreflightError(
            "Preflight FO reprovado; nenhuma alteração foi aplicada ao banco:\n" + details
        )


@dataclass(frozen=True)
class FoImportResult:
    processed_count: int
    updated_count: int
    inserted_count: int
    preserved_missing_count: int
    replaced_schedule_dates: bool
    preflight: FoImportPreflight


class FoImportPreflightError(ValueError):
    pass


FO_IMPORT_REQUIRED_LESSON_COLUMNS = {
    "slot_key",
    "lesson_code",
    "track_code",
    "lesson_type",
    "title_raw",
    "portal_title",
    "relative_path",
    "external_url",
    "duration_seconds",
    "subject_name",
    "subject_prefix",
    "module_label",
    "module_number",
    "lesson_number",
    "week_number",
    "day_index",
    "day_name",
    "slot_index",
    "recommended_date",
    "is_seen",
    "seen_at",
    "is_cut",
    "cut_reason",
    "cut_source",
    "source_sheet",
}
FO_IMPORT_REQUIRED_TABLE_COLUMNS = {
    "exercise_tasks": {
        "source_lesson_code",
        "scheduled_date",
        "status",
        "is_active",
        "manually_moved",
    },
    "app_settings": {"setting_key", "setting_value"},
    "daily_assignments": {
        "dashboard_date",
        "planned_slot_key",
        "assigned_lesson_code",
    },
}


def normalize_text(value: str) -> str:
    lowered = value.strip().lower()
    normalized = unicodedata.normalize("NFKD", lowered)
    return "".join(c for c in normalized if not unicodedata.combining(c))


def roman_to_int(roman: str) -> int:
    values = {"I": 1, "V": 5, "X": 10, "L": 50, "C": 100, "D": 500, "M": 1000}
    total = 0
    prev = 0
    for char in reversed(roman.upper()):
        value = values[char]
        if value < prev:
            total -= value
        else:
            total += value
            prev = value
    return total


def subject_prefix(subject_name: str) -> str:
    key = normalize_text(subject_name)
    if key in SUBJECT_PREFIXES:
        return SUBJECT_PREFIXES[key]

    letters = "".join(ch for ch in key if ch.isalpha())
    return (letters[:3] if letters else "UNK").upper()


def normalize_relative_path_key(value: str) -> str:
    return unicodedata.normalize("NFC", value.strip()).replace("\\", "/")


def load_un_duration_map(duration_csv_path: str | Path = UN_DURATION_CSV_PATH) -> dict[str, int]:
    path = Path(duration_csv_path)
    if not path.exists():
        return {}

    durations_by_path: dict[str, int] = {}
    with path.open(newline="", encoding="utf-8") as file_obj:
        reader = csv.DictReader(file_obj)
        for row in reader:
            relative_path = normalize_relative_path_key(
                (row.get("video_relative_path") or row.get("relative_path") or "").strip()
            )
            raw_duration = (row.get("duration_seconds") or "").strip()
            if not relative_path or not raw_duration:
                continue
            try:
                durations_by_path[relative_path] = int(round(float(raw_duration)))
            except ValueError:
                continue
    return durations_by_path


def parse_recommended_date(value: object) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    raise ValueError(f"Data inválida no cronograma: {value!r}")


def parse_subject_module(subject_module: str) -> tuple[str, str | None, int | None]:
    clean_value = subject_module.strip()
    module_label = None
    module_number = None
    subject_name = clean_value

    roman_match = ROMAN_RE.match(clean_value)
    if roman_match:
        subject_name = roman_match.group("subject").strip()
        module_label = roman_match.group("module").upper()
        module_number = roman_to_int(module_label)

    return subject_name, module_label, module_number


def parse_lesson_title(title: str) -> dict:
    clean_title = title.strip()

    if normalize_text(clean_title) == "revisao / livre":
        return {
            "lesson_type": "review",
            "subject_name": "Revisão",
            "subject_prefix": "REV",
            "module_label": None,
            "module_number": None,
            "lesson_number": None,
        }

    pending_match = PENDING_RE.match(clean_title)
    if pending_match:
        subject_name = pending_match.group("subject").strip()
        return {
            "lesson_type": "pending",
            "subject_name": subject_name,
            "subject_prefix": subject_prefix(subject_name),
            "module_label": None,
            "module_number": None,
            "lesson_number": None,
        }

    lesson_match = LESSON_RE.match(clean_title)
    if not lesson_match:
        raise ValueError(f"Formato de aula não reconhecido: {clean_title}")

    subject_module = lesson_match.group("subject_module").strip()
    lesson_number = int(lesson_match.group("lesson_number"))

    subject_name, module_label, module_number = parse_subject_module(subject_module)

    return {
        "lesson_type": "lesson",
        "subject_name": subject_name,
        "subject_prefix": subject_prefix(subject_name),
        "module_label": module_label,
        "module_number": module_number,
        "lesson_number": lesson_number,
    }


def parse_optional_int(value: object) -> int | None:
    text = str(value).strip() if value is not None else ""
    if not text:
        return None
    try:
        return int(round(float(text)))
    except ValueError:
        return None


def build_fo_metadata_key(
    subject_name: str | None,
    module_number: int | None,
    lesson_number: int | None,
) -> tuple[str, int | None, int] | None:
    if not subject_name or lesson_number is None:
        return None
    return (normalize_text(subject_name), module_number, lesson_number)


def choose_fo_metadata_row(current: dict | None, candidate: dict) -> dict:
    if current is None:
        return candidate

    current_score = (
        int(bool(current.get("duration_seconds")))
        + int(bool(current.get("external_url")))
        + int(bool(current.get("portal_title")))
    )
    candidate_score = (
        int(bool(candidate.get("duration_seconds")))
        + int(bool(candidate.get("external_url")))
        + int(bool(candidate.get("portal_title")))
    )
    return candidate if candidate_score > current_score else current


def load_fo_portal_metadata(csv_path: str | Path = FO_METADATA_CSV_PATH) -> dict[tuple[str, int | None, int], dict]:
    path = Path(csv_path)
    if not path.exists():
        return {}

    index: dict[tuple[str, int | None, int], dict] = {}

    with path.open(newline="", encoding="utf-8") as file_obj:
        reader = csv.DictReader(file_obj)
        for row in reader:
            subject_group = (row.get("subject_group") or "").strip()
            lesson_number = parse_optional_int(row.get("order_in_subject"))
            if not subject_group or lesson_number is None:
                continue

            group_parts = [part.strip() for part in subject_group.split(">") if part.strip()]
            subject_tail = group_parts[-1] if group_parts else subject_group
            subject_name, _, module_number = parse_subject_module(subject_tail)
            key = build_fo_metadata_key(subject_name, module_number, lesson_number)
            if key is None:
                continue

            candidate = {
                "portal_title": (row.get("lesson_title") or "").strip() or None,
                "duration_seconds": parse_optional_int(row.get("duration_seconds")),
                "external_url": (row.get("embed_url") or row.get("source_url") or "").strip() or None,
            }
            index[key] = choose_fo_metadata_row(index.get(key), candidate)

    return index


def build_lesson_code(parsed: dict, pending_counters: defaultdict[str, int], slot_key: str) -> str:
    lesson_type = parsed["lesson_type"]

    if lesson_type == "review":
        return f"REV{slot_key.replace('-', '')}"

    prefix = parsed["subject_prefix"]

    if lesson_type == "pending":
        pending_counters[prefix] += 1
        return f"{prefix}PEND{pending_counters[prefix]}"

    module_number = parsed["module_number"]
    module_part = str(module_number) if module_number is not None else ""
    lesson_number = parsed["lesson_number"]
    return f"{prefix}{module_part}A{lesson_number}"


def find_week_header_rows(sheet) -> list[tuple[int, int]]:
    rows: list[tuple[int, int]] = []
    for row in range(1, sheet.max_row + 1):
        value = sheet.cell(row, 2).value
        if isinstance(value, str):
            match = WEEK_HEADER_RE.match(value.strip())
            if match:
                rows.append((row, int(match.group(1))))
    return rows


def parse_sheet(workbook_path: str | Path, sheet_name: str) -> list[ParsedLesson]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Aba não encontrada: {sheet_name}")

    sheet = wb[sheet_name]
    week_headers = find_week_header_rows(sheet)
    if not week_headers:
        raise ValueError(f"Não foi encontrada nenhuma semana na aba {sheet_name}")

    parsed_lessons: list[ParsedLesson] = []
    pending_counters: defaultdict[str, int] = defaultdict(int)
    fo_metadata_index = load_fo_portal_metadata()

    for index, (header_row, week_number) in enumerate(week_headers):
        next_header_row = week_headers[index + 1][0] if index + 1 < len(week_headers) else sheet.max_row + 1
        date_row = header_row + 1

        slot_index = 0
        for row in range(header_row + 3, next_header_row):
            row_titles = []
            for col, _, _ in DAY_COLUMNS:
                value = sheet.cell(row, col).value
                if isinstance(value, str) and value.strip():
                    row_titles.append(value.strip())

            if not row_titles:
                continue

            slot_index += 1

            for col, day_index, day_name in DAY_COLUMNS:
                raw_value = sheet.cell(row, col).value
                if not isinstance(raw_value, str) or not raw_value.strip():
                    continue

                title_raw = raw_value.strip()
                parsed = parse_lesson_title(title_raw)

                recommended_date_value = sheet.cell(date_row, col).value
                recommended_date = parse_recommended_date(recommended_date_value)

                slot_key = f"W{week_number:02d}-D{day_index}-S{slot_index}"
                lesson_code = build_lesson_code(parsed, pending_counters, slot_key)
                metadata_key = build_fo_metadata_key(
                    subject_name=parsed["subject_name"],
                    module_number=parsed["module_number"],
                    lesson_number=parsed["lesson_number"],
                )
                fo_metadata = fo_metadata_index.get(metadata_key) if metadata_key else None

                parsed_lessons.append(
                    ParsedLesson(
                        slot_key=slot_key,
                        lesson_code=lesson_code,
                        track_code="FO",
                        lesson_type=parsed["lesson_type"],
                        title_raw=title_raw,
                        portal_title=fo_metadata.get("portal_title") if fo_metadata else None,
                        relative_path=None,
                        external_url=fo_metadata.get("external_url") if fo_metadata else None,
                        duration_seconds=fo_metadata.get("duration_seconds") if fo_metadata else None,
                        subject_name=parsed["subject_name"],
                        subject_prefix=parsed["subject_prefix"],
                        module_label=parsed["module_label"],
                        module_number=parsed["module_number"],
                        lesson_number=parsed["lesson_number"],
                        week_number=week_number,
                        day_index=day_index,
                        day_name=day_name,
                        slot_index=slot_index,
                        recommended_date=recommended_date,
                        source_sheet=sheet_name,
                    )
                )

    return parsed_lessons


def _duplicate_values(values: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    duplicates: set[str] = set()
    for value in values:
        if value in seen:
            duplicates.add(value)
        seen.add(value)
    return sorted(duplicates)


def build_empty_fo_import_preflight(lessons: Iterable[ParsedLesson]) -> FoImportPreflight:
    lesson_rows = list(lessons)
    errors: list[str] = []
    duplicate_codes = _duplicate_values(lesson.lesson_code for lesson in lesson_rows)
    duplicate_slots = _duplicate_values(lesson.slot_key for lesson in lesson_rows)
    if duplicate_codes:
        errors.append("lesson_code duplicado na planilha: " + ", ".join(duplicate_codes))
    if duplicate_slots:
        errors.append("slot_key duplicado na planilha: " + ", ".join(duplicate_slots))
    non_fo_rows = [lesson.lesson_code for lesson in lesson_rows if lesson.track_code != "FO"]
    if non_fo_rows:
        errors.append("itens não FO recebidos pelo import FO: " + ", ".join(sorted(non_fo_rows)))

    return FoImportPreflight(
        incoming_count=len(lesson_rows),
        existing_fo_count=0,
        matching_lesson_codes=(),
        new_lesson_codes=tuple(sorted({lesson.lesson_code for lesson in lesson_rows})),
        missing_lesson_codes=(),
        errors=tuple(errors),
    )


def build_fo_import_preflight(
    conn: sqlite3.Connection,
    lessons: Iterable[ParsedLesson],
) -> FoImportPreflight:
    lesson_rows = list(lessons)
    base_report = build_empty_fo_import_preflight(lesson_rows)
    errors = list(base_report.errors)

    table_row = conn.execute(
        "SELECT 1 FROM sqlite_master WHERE type = 'table' AND name = 'lessons'"
    ).fetchone()
    if not table_row:
        errors.append("tabela lessons ausente; inicialize um banco novo antes do import")
        return FoImportPreflight(
            incoming_count=base_report.incoming_count,
            existing_fo_count=0,
            matching_lesson_codes=(),
            new_lesson_codes=base_report.new_lesson_codes,
            missing_lesson_codes=(),
            errors=tuple(errors),
        )

    lesson_columns = {
        row["name"] for row in conn.execute("PRAGMA table_info(lessons)").fetchall()
    }
    missing_columns = sorted(FO_IMPORT_REQUIRED_LESSON_COLUMNS - lesson_columns)
    if missing_columns:
        errors.append(
            "schema de lessons incompatível; colunas ausentes: " + ", ".join(missing_columns)
        )
        return FoImportPreflight(
            incoming_count=base_report.incoming_count,
            existing_fo_count=0,
            matching_lesson_codes=(),
            new_lesson_codes=base_report.new_lesson_codes,
            missing_lesson_codes=(),
            errors=tuple(errors),
        )

    for table_name, required_columns in FO_IMPORT_REQUIRED_TABLE_COLUMNS.items():
        table_exists = conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type = 'table' AND name = ?",
            (table_name,),
        ).fetchone()
        if not table_exists:
            errors.append(f"tabela obrigatória ausente: {table_name}")
            continue
        table_columns = {
            row["name"]
            for row in conn.execute(f'PRAGMA table_info("{table_name}")').fetchall()
        }
        missing_table_columns = sorted(required_columns - table_columns)
        if missing_table_columns:
            errors.append(
                f"schema de {table_name} incompatível; colunas ausentes: "
                + ", ".join(missing_table_columns)
            )

    existing_rows = conn.execute(
        "SELECT slot_key, lesson_code, track_code, lesson_type FROM lessons"
    ).fetchall()
    duplicate_existing_codes = _duplicate_values(row["lesson_code"] for row in existing_rows)
    duplicate_existing_slots = _duplicate_values(row["slot_key"] for row in existing_rows)
    if duplicate_existing_codes:
        errors.append("lesson_code duplicado no banco: " + ", ".join(duplicate_existing_codes))
    if duplicate_existing_slots:
        errors.append("slot_key duplicado no banco: " + ", ".join(duplicate_existing_slots))
    existing_by_code = {row["lesson_code"]: row for row in existing_rows}
    existing_by_slot = {row["slot_key"]: row for row in existing_rows}
    existing_fo_codes = {
        row["lesson_code"] for row in existing_rows if row["track_code"] == "FO"
    }

    matching_codes: list[str] = []
    new_codes: list[str] = []
    incoming_codes = {lesson.lesson_code for lesson in lesson_rows}

    if not base_report.errors:
        for lesson in lesson_rows:
            code_row = existing_by_code.get(lesson.lesson_code)
            slot_row = existing_by_slot.get(lesson.slot_key)

            if code_row is None and slot_row is None:
                new_codes.append(lesson.lesson_code)
                continue

            same_identity = (
                code_row is not None
                and slot_row is not None
                and code_row["lesson_code"] == lesson.lesson_code
                and slot_row["lesson_code"] == lesson.lesson_code
                and code_row["slot_key"] == lesson.slot_key
                and slot_row["slot_key"] == lesson.slot_key
                and code_row["track_code"] == "FO"
            )
            if same_identity:
                if code_row["lesson_type"] != lesson.lesson_type:
                    errors.append(
                        f"tipo alterado para {lesson.lesson_code} no slot {lesson.slot_key}: "
                        f"banco={code_row['lesson_type']} planilha={lesson.lesson_type}"
                    )
                else:
                    matching_codes.append(lesson.lesson_code)
                continue

            if code_row is not None and code_row["slot_key"] != lesson.slot_key:
                errors.append(
                    f"aula movida: {lesson.lesson_code} está em {code_row['slot_key']} no banco "
                    f"e em {lesson.slot_key} na planilha"
                )
            if slot_row is not None and slot_row["lesson_code"] != lesson.lesson_code:
                errors.append(
                    f"slot ocupado: {lesson.slot_key} contém {slot_row['lesson_code']} no banco "
                    f"e receberia {lesson.lesson_code}"
                )
            if code_row is not None and code_row["track_code"] != "FO":
                errors.append(
                    f"colisão entre trilhas: {lesson.lesson_code} já pertence a {code_row['track_code']}"
                )
            if code_row is None and slot_row is not None and slot_row["lesson_code"] == lesson.lesson_code:
                errors.append(
                    f"identidade ambígua no slot {lesson.slot_key} para {lesson.lesson_code}"
                )

    return FoImportPreflight(
        incoming_count=len(lesson_rows),
        existing_fo_count=len(existing_fo_codes),
        matching_lesson_codes=tuple(sorted(matching_codes)),
        new_lesson_codes=tuple(sorted(new_codes)),
        missing_lesson_codes=tuple(sorted(existing_fo_codes - incoming_codes)),
        errors=tuple(errors),
    )


def format_fo_import_preflight(report: FoImportPreflight) -> str:
    def format_codes(codes: tuple[str, ...], limit: int = 20) -> str:
        visible = codes[:limit]
        suffix = f" ... (+{len(codes) - limit})" if len(codes) > limit else ""
        return ", ".join(visible) + suffix

    lines = [
        "Preflight FO:",
        f"- itens na planilha: {report.incoming_count}",
        f"- itens FO existentes: {report.existing_fo_count}",
        f"- identidades coincidentes: {len(report.matching_lesson_codes)}",
        f"- aulas novas: {len(report.new_lesson_codes)}",
        f"- aulas ausentes preservadas: {len(report.missing_lesson_codes)}",
        f"- conflitos: {len(report.errors)}",
    ]
    if report.new_lesson_codes:
        lines.append("- novos códigos: " + format_codes(report.new_lesson_codes))
    if report.missing_lesson_codes:
        lines.append("- códigos ausentes preservados: " + format_codes(report.missing_lesson_codes))
    if report.errors:
        lines.append("Conflitos detectados:")
        lines.extend(f"- {message}" for message in report.errors)
    return "\n".join(lines)


def lesson_hash(value: str) -> str:
    return hashlib.sha1(value.encode("utf-8")).hexdigest()[:12].upper()


def build_un_lesson_code(subject_prefix_value: str, relative_path: str) -> str:
    return f"UN{subject_prefix_value}-{lesson_hash(relative_path)}"


def week_number_from_start(start_date: date, current_date_value: date) -> int:
    return ((current_date_value - start_date).days // 7) + 1


def business_days_between(start_date: date, end_date: date) -> list[date]:
    days: list[date] = []
    current = start_date
    while current <= end_date:
        if current.isoweekday() <= 5:
            days.append(current)
        current += timedelta(days=1)
    return days


def distribute_items_across_days(items: list[dict], schedule_days: list[date]) -> list[tuple[date, dict]]:
    if not items:
        return []
    if not schedule_days:
        raise ValueError("Não há dias úteis restantes para distribuir as aulas do Universo Narrado.")

    base, remainder = divmod(len(items), len(schedule_days))
    allocations = [base + (1 if index < remainder else 0) for index in range(len(schedule_days))]

    distribution: list[tuple[date, dict]] = []
    item_index = 0
    for schedule_day, allocation in zip(schedule_days, allocations):
        for _ in range(allocation):
            distribution.append((schedule_day, items[item_index]))
            item_index += 1
    return distribution


def parse_universe_narrado_csv(
    csv_path: str | Path,
    start_date: date,
    end_date: date,
) -> list[ParsedLesson]:
    path = Path(csv_path)
    if not path.exists():
        raise ValueError(f"CSV do Universo Narrado não encontrado: {path}")

    schedule_days = business_days_between(start_date=start_date, end_date=end_date)
    grouped_rows: dict[str, list[dict]] = {"FIS": [], "MAT": []}
    durations_by_path = load_un_duration_map()

    with path.open(newline="", encoding="utf-8") as file_obj:
        reader = csv.DictReader(file_obj)
        for row in reader:
            subject = (row.get("subject") or "").strip().upper()
            if subject not in grouped_rows:
                continue
            if (row.get("is_active") or "").strip() != "1":
                continue
            grouped_rows[subject].append(row)

    distributions: list[tuple[date, dict]] = []
    for subject in ("FIS", "MAT"):
        distributions.extend(distribute_items_across_days(grouped_rows[subject], schedule_days))

    distributions.sort(
        key=lambda item: (
            item[0],
            item[1]["subject"],
            int(item[1].get("sequence_order") or item[1].get("lesson_order") or 0),
            item[1].get("lesson_title") or "",
        )
    )

    slot_counters: defaultdict[str, int] = defaultdict(int)
    parsed_lessons: list[ParsedLesson] = []

    for recommended_date_value, row in distributions:
        recommended_date = recommended_date_value.isoformat()
        slot_counters[recommended_date] += 1

        subject = row["subject"].strip().upper()
        topic = (row.get("topic") or "").strip()
        subtopic = (row.get("subtopic") or "").strip()
        lesson_title = (row.get("lesson_title") or "").strip()
        relative_path = (row.get("relative_path") or "").strip()
        item_type = (row.get("item_type") or "lesson").strip().lower()
        module_label = topic if subtopic in {"", "Sem Tópico"} else f"{topic} | {subtopic}"
        day_index = recommended_date_value.isoweekday()
        raw_lesson_order = (row.get("lesson_order") or "").strip()
        lesson_number = int(raw_lesson_order) if item_type == "lesson" and raw_lesson_order.isdigit() else None
        raw_duration = (row.get("duration_seconds") or "").strip()
        if item_type == "lesson" and raw_duration:
            duration_seconds = int(round(float(raw_duration)))
        elif item_type == "lesson" and relative_path:
            duration_seconds = durations_by_path.get(normalize_relative_path_key(relative_path))
        else:
            duration_seconds = None

        parsed_lessons.append(
            ParsedLesson(
                slot_key=build_un_lesson_code(subject, relative_path),
                lesson_code=build_un_lesson_code(subject, relative_path),
                track_code="UN",
                lesson_type=item_type,
                title_raw=lesson_title,
                portal_title=None,
                relative_path=relative_path or None,
                external_url=None,
                duration_seconds=duration_seconds,
                subject_name="Fisica" if subject == "FIS" else "Matematica",
                subject_prefix=subject,
                module_label=module_label,
                module_number=None,
                lesson_number=lesson_number,
                week_number=week_number_from_start(start_date=start_date, current_date_value=recommended_date_value),
                day_index=day_index,
                day_name=PORTUGUESE_DAY_NAMES[day_index],
                slot_index=slot_counters[recommended_date],
                recommended_date=recommended_date,
                source_sheet="universo_narrado_csv",
            )
        )

    return parsed_lessons


def upsert_lessons(
    conn: sqlite3.Connection,
    lessons: Iterable[ParsedLesson],
    source_sheet: str,
    clear_daily_assignments: bool = False,
    prune_missing: bool = True,
) -> int:
    lesson_rows = list(lessons)
    if not lesson_rows:
        return 0

    with conn:
        conn.executemany(
            """
            INSERT INTO lessons (
                slot_key,
                lesson_code,
                track_code,
                lesson_type,
                title_raw,
                portal_title,
                relative_path,
                external_url,
                duration_seconds,
                subject_name,
                subject_prefix,
                module_label,
                module_number,
                lesson_number,
                week_number,
                day_index,
                day_name,
                slot_index,
                recommended_date,
                source_sheet,
                updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(slot_key) DO UPDATE SET
                lesson_code = excluded.lesson_code,
                track_code = excluded.track_code,
                lesson_type = excluded.lesson_type,
                title_raw = excluded.title_raw,
                portal_title = COALESCE(excluded.portal_title, lessons.portal_title),
                relative_path = excluded.relative_path,
                external_url = COALESCE(excluded.external_url, lessons.external_url),
                duration_seconds = COALESCE(excluded.duration_seconds, lessons.duration_seconds),
                subject_name = excluded.subject_name,
                subject_prefix = excluded.subject_prefix,
                module_label = excluded.module_label,
                module_number = excluded.module_number,
                lesson_number = excluded.lesson_number,
                week_number = excluded.week_number,
                day_index = excluded.day_index,
                day_name = excluded.day_name,
                slot_index = excluded.slot_index,
                recommended_date = excluded.recommended_date,
                source_sheet = excluded.source_sheet,
                updated_at = CURRENT_TIMESTAMP;
            """,
            [
                (
                    lesson.slot_key,
                    lesson.lesson_code,
                    lesson.track_code,
                    lesson.lesson_type,
                    lesson.title_raw,
                    lesson.portal_title,
                    lesson.relative_path,
                    lesson.external_url,
                    lesson.duration_seconds,
                    lesson.subject_name,
                    lesson.subject_prefix,
                    lesson.module_label,
                    lesson.module_number,
                    lesson.lesson_number,
                    lesson.week_number,
                    lesson.day_index,
                    lesson.day_name,
                    lesson.slot_index,
                    lesson.recommended_date,
                    lesson.source_sheet,
                )
                for lesson in lesson_rows
            ],
        )
        conn.executemany(
            """
            UPDATE lessons
            SET duration_seconds = ?,
                external_url = COALESCE(?, external_url),
                portal_title = COALESCE(?, portal_title),
                updated_at = CURRENT_TIMESTAMP
            WHERE slot_key = ?
            """,
            [
                (lesson.duration_seconds, lesson.external_url, lesson.portal_title, lesson.slot_key)
                for lesson in lesson_rows
                if lesson.duration_seconds is not None or lesson.external_url is not None or lesson.portal_title
            ],
        )

        if prune_missing:
            imported_slot_keys = [lesson.slot_key for lesson in lesson_rows]
            conn.execute("DROP TABLE IF EXISTS imported_slots")
            conn.execute("CREATE TEMP TABLE imported_slots (slot_key TEXT PRIMARY KEY)")
            conn.executemany(
                "INSERT INTO imported_slots (slot_key) VALUES (?)",
                [(slot_key,) for slot_key in imported_slot_keys],
            )
            conn.execute(
                """
                DELETE FROM lessons
                WHERE source_sheet = ?
                  AND slot_key NOT IN (SELECT slot_key FROM imported_slots)
                """,
                (source_sheet,),
            )
            conn.execute("DROP TABLE imported_slots")
        conn.execute(
            """
            DELETE FROM daily_assignments
            WHERE assigned_lesson_code NOT IN (SELECT lesson_code FROM lessons)
               OR planned_slot_key NOT IN (SELECT slot_key FROM lessons)
            """
        )
        conn.execute(
            """
            DELETE FROM un_daily_assignments
            WHERE assigned_lesson_code NOT IN (SELECT lesson_code FROM lessons)
            """
        )
        if clear_daily_assignments:
            conn.execute("DELETE FROM daily_assignments")

    return len(lesson_rows)


def import_fo_lessons(
    conn: sqlite3.Connection,
    lessons: Iterable[ParsedLesson],
    *,
    replace_schedule_dates: bool = False,
    preflight_reporter: Callable[[FoImportPreflight], None] | None = None,
) -> FoImportResult:
    lesson_rows = list(lessons)
    preflight = build_fo_import_preflight(conn=conn, lessons=lesson_rows)
    if preflight_reporter:
        preflight_reporter(preflight)
    preflight.require_safe()

    if conn.in_transaction:
        raise RuntimeError("O import FO exige uma conexão sem transação pendente.")

    lessons_by_code = {lesson.lesson_code: lesson for lesson in lesson_rows}
    matching_lessons = [
        lessons_by_code[lesson_code] for lesson_code in preflight.matching_lesson_codes
    ]
    new_lessons = [
        lessons_by_code[lesson_code] for lesson_code in preflight.new_lesson_codes
    ]

    try:
        conn.execute("BEGIN IMMEDIATE")
        locked_preflight = build_fo_import_preflight(conn=conn, lessons=lesson_rows)
        if locked_preflight != preflight:
            raise FoImportPreflightError(
                "O banco mudou depois do preflight; nenhuma alteração foi aplicada. Rode o import novamente."
            )

        if replace_schedule_dates:
            conn.executemany(
                """
                UPDATE lessons
                SET lesson_type = ?,
                    title_raw = ?,
                    portal_title = COALESCE(?, portal_title),
                    relative_path = ?,
                    external_url = COALESCE(?, external_url),
                    duration_seconds = COALESCE(?, duration_seconds),
                    subject_name = ?,
                    subject_prefix = ?,
                    module_label = ?,
                    module_number = ?,
                    lesson_number = ?,
                    week_number = ?,
                    day_index = ?,
                    day_name = ?,
                    slot_index = ?,
                    recommended_date = ?,
                    source_sheet = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE lesson_code = ?
                  AND slot_key = ?
                  AND track_code = 'FO'
                """,
                [
                    (
                        lesson.lesson_type,
                        lesson.title_raw,
                        lesson.portal_title,
                        lesson.relative_path,
                        lesson.external_url,
                        lesson.duration_seconds,
                        lesson.subject_name,
                        lesson.subject_prefix,
                        lesson.module_label,
                        lesson.module_number,
                        lesson.lesson_number,
                        lesson.week_number,
                        lesson.day_index,
                        lesson.day_name,
                        lesson.slot_index,
                        lesson.recommended_date,
                        lesson.source_sheet,
                        lesson.lesson_code,
                        lesson.slot_key,
                    )
                    for lesson in matching_lessons
                ],
            )
        else:
            conn.executemany(
                """
                UPDATE lessons
                SET lesson_type = ?,
                    title_raw = ?,
                    portal_title = COALESCE(?, portal_title),
                    relative_path = ?,
                    external_url = COALESCE(?, external_url),
                    duration_seconds = COALESCE(?, duration_seconds),
                    subject_name = ?,
                    subject_prefix = ?,
                    module_label = ?,
                    module_number = ?,
                    lesson_number = ?,
                    source_sheet = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE lesson_code = ?
                  AND slot_key = ?
                  AND track_code = 'FO'
                """,
                [
                    (
                        lesson.lesson_type,
                        lesson.title_raw,
                        lesson.portal_title,
                        lesson.relative_path,
                        lesson.external_url,
                        lesson.duration_seconds,
                        lesson.subject_name,
                        lesson.subject_prefix,
                        lesson.module_label,
                        lesson.module_number,
                        lesson.lesson_number,
                        lesson.source_sheet,
                        lesson.lesson_code,
                        lesson.slot_key,
                    )
                    for lesson in matching_lessons
                ],
            )

        conn.executemany(
            """
            INSERT INTO lessons (
                slot_key,
                lesson_code,
                track_code,
                lesson_type,
                title_raw,
                portal_title,
                relative_path,
                external_url,
                duration_seconds,
                subject_name,
                subject_prefix,
                module_label,
                module_number,
                lesson_number,
                week_number,
                day_index,
                day_name,
                slot_index,
                recommended_date,
                source_sheet,
                updated_at
            ) VALUES (?, ?, 'FO', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            """,
            [
                (
                    lesson.slot_key,
                    lesson.lesson_code,
                    lesson.lesson_type,
                    lesson.title_raw,
                    lesson.portal_title,
                    lesson.relative_path,
                    lesson.external_url,
                    lesson.duration_seconds,
                    lesson.subject_name,
                    lesson.subject_prefix,
                    lesson.module_label,
                    lesson.module_number,
                    lesson.lesson_number,
                    lesson.week_number,
                    lesson.day_index,
                    lesson.day_name,
                    lesson.slot_index,
                    lesson.recommended_date,
                    lesson.source_sheet,
                )
                for lesson in new_lessons
            ],
        )

        new_video_lessons = [
            lesson for lesson in new_lessons if lesson.lesson_type == "lesson"
        ]
        if new_video_lessons:
            offset_days = get_exercise_offset_days(conn)
            conn.executemany(
                """
                INSERT INTO exercise_tasks (
                    source_lesson_code,
                    scheduled_date,
                    status,
                    is_active,
                    manually_moved,
                    updated_at
                ) VALUES (?, ?, 'pending', 0, 0, CURRENT_TIMESTAMP)
                ON CONFLICT(source_lesson_code) DO NOTHING
                """,
                [
                    (
                        lesson.lesson_code,
                        scheduled_date_from(date.fromisoformat(lesson.recommended_date), offset_days),
                    )
                    for lesson in new_video_lessons
                ],
            )

        if replace_schedule_dates:
            conn.execute("DELETE FROM daily_assignments")

        conn.commit()
    except Exception:
        conn.rollback()
        raise

    return FoImportResult(
        processed_count=len(lesson_rows),
        updated_count=len(matching_lessons),
        inserted_count=len(new_lessons),
        preserved_missing_count=len(preflight.missing_lesson_codes),
        replaced_schedule_dates=replace_schedule_dates,
        preflight=preflight,
    )


def import_workbook_sheet(
    conn: sqlite3.Connection,
    workbook_path: str | Path,
    sheet_name: str,
    *,
    replace_schedule_dates: bool = False,
    preflight_reporter: Callable[[FoImportPreflight], None] | None = None,
) -> int:
    parsed = parse_sheet(workbook_path=workbook_path, sheet_name=sheet_name)
    result = import_fo_lessons(
        conn=conn,
        lessons=parsed,
        replace_schedule_dates=replace_schedule_dates,
        preflight_reporter=preflight_reporter,
    )
    return result.processed_count


def universe_narrado_end_date(conn: sqlite3.Connection) -> date:
    row = conn.execute(
        """
        SELECT MAX(recommended_date) AS end_date
        FROM lessons
        WHERE track_code = 'FO'
          AND lesson_type = 'lesson'
        """
    ).fetchone()
    end_date = row["end_date"] if row else None
    if not end_date:
        raise ValueError("Importe o cronograma FO antes de importar o Universo Narrado.")
    return date.fromisoformat(end_date)


def import_universe_narrado_csv(
    conn: sqlite3.Connection,
    csv_path: str | Path,
    start_date: date | None = None,
) -> int:
    effective_start_date = start_date or DEFAULT_UN_START_DATE
    end_date = universe_narrado_end_date(conn=conn)
    parsed = parse_universe_narrado_csv(
        csv_path=csv_path,
        start_date=effective_start_date,
        end_date=end_date,
    )
    imported_count = upsert_lessons(
        conn=conn,
        lessons=parsed,
        source_sheet="universo_narrado_csv",
        clear_daily_assignments=False,
        prune_missing=False,
    )
    with conn:
        conn.execute("DELETE FROM un_daily_assignments")
    return imported_count
